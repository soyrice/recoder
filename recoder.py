# import arcpy site packages and support modules
# quit script if required modules aren't installed
try:
    import arcpy
    import openpyxl
    import re

except:
    arcpy.AddMessage("Required modules aren't installed. Please install openpyxl.")
    exit()

def main() :
    # Read the parameter values
    # 1. name = Excel worksheet, type = any value
    # 2. name = Excel column name, type = string
    # 3. name = Input workspace, type = workspace
    # 4. name = Input features, type = feature class
    # 5. name = Field name, type = field, dependency = Input features
    #
    worksheetPath = arcpy.GetParameterAsText(0)
    columnHeader = arcpy.GetParameterAsText(1)
    arcpy.env.workspace = arcpy.GetParameterAsText(2)
    data = arcpy.GetParameterAsText(3)
    LULCField = arcpy.GetParameterAsText(4)

    # parse the worksheet path string to find the file path (path up to .xlsx)
    # and the worksheet name (path b/w quotes after .xlsx)
    filePath = re.findall("(?<=)(.*?)(?=')", worksheetPath)[0][:-1]
    worksheetName = re.findall("(?<=')(.*)(?=\$)", worksheetPath)[0]

    generalizedLULCDict = importData(filePath, worksheetName, columnHeader)
    field_names = addFields(generalizedLULCDict, data, LULCField)
    recodeFLUM(field_names, generalizedLULCDict, data, LULCField)

def importData(filePath, worksheetName, columnHeader):
    arcpy.AddMessage('Getting data from Excel sheet...')

    # create new variables for each generalized land use class
    wb = openpyxl.load_workbook(filePath)
    sheet = wb.get_sheet_by_name(worksheetName)

    # map each possible muncipal LULC class under columnHeader to each generalized LULC
    # because each key will be a field name in ArcMap, it can only be 10 characters long
    genList, lulcList = [], []
    for row in sheet.iter_rows() :
        for cell in row :
            if cell.value == columnHeader :
                for r in range(2,sheet.max_row):
                    lulc = str(sheet['%s%s'%(re.findall("(.*)[0-9]", cell.coordinate)[0], r)].value).upper()
                    lulcList.append(re.findall('[^;]+',re.sub(' ', '', lulc)))
        if row[0].value != None and row[0].value != 'Generalized LULC Class' :
            genList.append(str(re.sub(r'\W+', '', row[0].value)[0:10]).upper())
    generalizedLULCDict = dict(zip(genList, lulcList))

    return generalizedLULCDict

def addFields(generalizedLULCDict, data, LULCField) :
    arcpy.AddMessage('Adding fields to feature class...')
    
    # add field for each generalized LULC class
    # save fields for da cursor
    field_names = []
    for key in generalizedLULCDict.keys() :
        arcpy.AddField_management(data, key, "SHORT")
        field_names.append(key)
        
    # add field to flag instances where no match exists (no information is added)
    arcpy.AddField_management(data, "NoMatch", "SHORT")
    field_names.append("NoMatch")

    # update generalized LULC class if muncipal LULC class is in that generalized category
    # if all matches succeed, each row should have a 1 value attributed to a single
    # generalized variable, and Null values for all other generalized variables
    field_names.append(LULCField)

    return field_names

def recodeFLUM(field_names, generalizedLULCDict, data, LULCField) :
    with arcpy.da.UpdateCursor(data, tuple(field_names)) as cursor :
        arcpy.AddMessage('Recoding records...')
                                                                           
        for row in cursor :
            fieldValue = row[-1]

            # match municpal land use class to generalized land use class
            # update the corresponding generalized land use class variable with a 1
            counter = 0
            for field in field_names[:-2] :

                if re.sub(' ', '',fieldValue.upper()) in generalizedLULCDict[field] :
                    row[counter] = 1
                    break
                
                # flag instances where no match is found
                if counter+1 == len(field_names)-2 :
                        row[-2] = 1
                counter += 1
                
            cursor.updateRow(row)

main()
